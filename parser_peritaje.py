# -*- coding: utf-8 -*-
"""
Parser local del peritaje Excel (formato Agrosalta / Cooperación).

Lee la grilla del Excel y extrae directamente las piezas marcadas con X
y sus precios, sin pasar por Claude. Esto es rápido, determinístico y
no inventa piezas inexistentes.

Estructura típica del peritaje:
- 3 sectores en paralelo por bloque de filas: columnas A-D, E-I, J-N.
- Cada sector empieza con un header (ej "PARTE DELANTERA") seguido por
  la fila de encabezado de columnas (A=Cambiar, B=Reparar, Precio).
- Las piezas van debajo del header. Una "X" sola en la columna A o B
  indica acción CAMBIAR o REPARAR respectivamente.
"""

from io import BytesIO


# Sectores conocidos y sufijo a agregar al nombre de la pieza
# para evitar ambigüedad (ej: "Guardabarro izquierdo" del sector
# "PARTE DELANTERA" → "Guardabarro delantero izquierdo").
SECTORES = {
    "PARTE DELANTERA": "delantero",
    "PARTE TRASERA": "trasero",
    "LADO IZQUIERDO": "izquierdo",
    "LADO DERECHO": "derecho",
    "MOTOR": "motor",
    "TREN DELANTERO": "tren delantero",
    "TREN TRASERO": "tren trasero",
    "PARTE INTERIOR": "interior",
    "CHASIS": "chasis",
    "OTROS": "",
}

# Posiciones típicas donde arrancan los 3 sectores paralelos:
# columna 0 (A), columna 4 (E), columna 9 (J). Para cada una, el offset
# de la columna de "Cambiar" (X), "Reparar" (X) y "Precio" relativos al
# inicio del sector.
LAYOUTS_SECTORES = {
    0:  (1, 2, 3),   # A=nombre, B=X-Cambiar, C=X-Reparar, D=Precio
    4:  (2, 3, 4),   # E=nombre, F=código, G=X-Cambiar, H=X-Reparar, I=Precio  (codigo en F se ignora)
    9:  (2, 3, 4),   # J=nombre, K=código, L=X-Cambiar, M=X-Reparar, N=Precio
}


def _es_x(valor):
    """True si la celda contiene UNICAMENTE la letra X (sola)."""
    if valor is None:
        return False
    s = str(valor).strip()
    return s.upper() == "X"


def _a_numero(valor):
    """Convierte un valor (str o numérico) a int. Devuelve 0 si no es número."""
    if valor is None or valor == "":
        return 0
    if isinstance(valor, (int, float)):
        return int(valor)
    s = str(valor).strip().replace("$", "").replace(" ", "")
    # Formato AR: "1.708.000,50" -> 1708000.50
    if "," in s:
        s = s.replace(".", "").replace(",", ".")
    elif s.count(".") >= 2:
        # "1.708.000" -> "1708000"
        s = s.replace(".", "")
    elif s.count(".") == 1:
        # "1.708" en AR es probablemente miles, no decimal.
        # Si los dígitos después del punto son exactamente 3, asumir miles.
        partes = s.split(".")
        if len(partes[1]) == 3:
            s = partes[0] + partes[1]
    try:
        return int(float(s))
    except (ValueError, TypeError):
        return 0


def _normalizar_sector(texto):
    """Normaliza el texto de una celda y devuelve el sector si es uno
    conocido, o None si no lo es."""
    if not isinstance(texto, str):
        return None
    s = texto.strip().upper()
    if s in SECTORES:
        return s
    return None


def _enriquecer_nombre(nombre_pieza, sufijo_sector):
    """Combina el nombre de la pieza con el sufijo del sector,
    evitando duplicar si ya aparece en el nombre."""
    nombre = str(nombre_pieza).strip()
    if not sufijo_sector:
        return nombre
    # Si el sufijo ya aparece en el nombre (caso "Paragolpe delantero"
    # en sector "PARTE DELANTERA"), no duplicar.
    nombre_lower = nombre.lower()
    sufijo_lower = sufijo_sector.lower()
    if sufijo_lower in nombre_lower:
        return nombre
    # Algunos sufijos en singular masculino tienen forma femenina:
    # "delantero" → si el nombre ya tiene "delantera", tampoco duplicar.
    pares = {
        "delantero": "delantera",
        "trasero": "trasera",
        "izquierdo": "izquierda",
        "derecho": "derecha",
    }
    alt = pares.get(sufijo_lower)
    if alt and alt in nombre_lower:
        return nombre
    return f"{nombre} {sufijo_sector}"


def _leer_grilla(datos, nombre_archivo):
    """
    Lee el Excel a una matriz de celdas: dict {(fila, col): valor}.
    Soporta .xls (xlrd) y .xlsx (openpyxl). Devuelve también el número
    de filas y columnas de la hoja principal.
    """
    nombre = nombre_archivo.lower()
    # Detección por contenido por si la extensión miente.
    es_xls_viejo = datos.startswith(b"\xd0\xcf\x11\xe0")
    es_xlsx = datos.startswith(b"PK")
    if not (es_xls_viejo or es_xlsx):
        # Decidir por extensión como fallback.
        es_xls_viejo = nombre.endswith(".xls") and not nombre.endswith(".xlsx")
        es_xlsx = nombre.endswith((".xlsx", ".xlsm"))

    celdas = {}
    nrows = ncols = 0

    if es_xls_viejo:
        import xlrd
        wb = xlrd.open_workbook(file_contents=datos)
        sh = wb.sheets()[0]
        nrows, ncols = sh.nrows, sh.ncols
        for r in range(nrows):
            for c in range(ncols):
                v = sh.cell_value(r, c)
                if v not in ("", None):
                    celdas[(r, c)] = v
    elif es_xlsx:
        from openpyxl import load_workbook
        wb = load_workbook(BytesIO(datos), data_only=True)
        if not wb.sheetnames:
            return {}, 0, 0
        ws = wb.worksheets[0]
        nrows, ncols = ws.max_row or 0, ws.max_column or 0
        for fila in ws.iter_rows():
            for celda in fila:
                if celda.value not in (None, ""):
                    celdas[(celda.row - 1, celda.column - 1)] = celda.value
    return celdas, nrows, ncols


def _detectar_sectores(celdas):
    """
    Encuentra todas las posiciones (fila, col, sector) donde hay un
    header de sector. Devuelve una lista ordenada.
    """
    sectores_encontrados = []
    for (r, c), v in celdas.items():
        sector = _normalizar_sector(v)
        if sector and c in LAYOUTS_SECTORES:
            sectores_encontrados.append((r, c, sector))
    sectores_encontrados.sort()
    return sectores_encontrados


def _piezas_de_sector(celdas, nrows, fila_inicio, col_inicio, sector,
                      siguientes_sectores_misma_col):
    """
    Extrae las piezas de un sector dado.
    - fila_inicio: fila donde está el header del sector.
    - col_inicio: columna donde está el nombre de la pieza.
    - sector: nombre del sector (para enriquecer nombres).
    - siguientes_sectores_misma_col: lista de filas donde arrancan
      sectores siguientes en la MISMA columna, para saber dónde
      detenernos.
    """
    sufijo = SECTORES.get(sector, "")
    offsets = LAYOUTS_SECTORES[col_inicio]
    col_cambiar = col_inicio + offsets[0]
    col_reparar = col_inicio + offsets[1]
    col_precio = col_inicio + offsets[2]

    # Determinar la fila de fin: la próxima vez que aparezca un sector
    # en esta misma columna, o el final de la grilla.
    fila_fin = nrows
    for f in siguientes_sectores_misma_col:
        if f > fila_inicio:
            fila_fin = f
            break

    piezas = []
    # Empezar a buscar piezas 1-3 filas después del header (puede haber
    # una fila intermedia con "A | B | Precio"). Detectamos esa fila
    # automáticamente saltando filas donde el "nombre" sea exactamente
    # "A" o vacío.
    for r in range(fila_inicio + 1, fila_fin):
        nombre = celdas.get((r, col_inicio))
        # Saltar la fila de encabezado "A | B | Precio" o equivalente
        if isinstance(nombre, str) and nombre.strip().upper() in ("A", "B", "PRECIO"):
            continue
        if not nombre:
            continue
        marca_cambiar = celdas.get((r, col_cambiar))
        marca_reparar = celdas.get((r, col_reparar))
        accion = None
        if _es_x(marca_cambiar):
            accion = "CAMBIAR"
        elif _es_x(marca_reparar):
            accion = "REPARAR"
        if not accion:
            continue
        precio = _a_numero(celdas.get((r, col_precio))) if accion == "CAMBIAR" else 0
        nombre_final = _enriquecer_nombre(nombre, sufijo)
        piezas.append({
            "accion": accion,
            "pieza": nombre_final,
            "precio": precio,
        })
    return piezas


def parsear_peritaje_excel(datos, nombre_archivo):
    """
    Parser principal. Recibe los bytes del Excel y devuelve un dict
    con las claves "danos" (lista de piezas) y "totalRepuestos" (int).

    Si el archivo no se puede leer o no se detectan sectores conocidos,
    devuelve None para que el flujo caiga al método anterior (Claude).
    """
    try:
        celdas, nrows, ncols = _leer_grilla(datos, nombre_archivo)
    except Exception:
        return None

    if not celdas:
        return None

    sectores = _detectar_sectores(celdas)
    if not sectores:
        return None

    # Agrupar sectores por columna para saber dónde termina cada uno.
    por_columna = {}
    for r, c, sector in sectores:
        por_columna.setdefault(c, []).append(r)
    for c in por_columna:
        por_columna[c].sort()

    danos = []
    for r, c, sector in sectores:
        sig = [f for f in por_columna[c] if f > r]
        piezas = _piezas_de_sector(celdas, nrows, r, c, sector, sig)
        danos.extend(piezas)

    resultado = {"danos": danos}

    # Extraer los items de mano de obra tal cual vienen (respetando lo
    # que el perito escribió, ej: "MECANICA Y TAPICERIA" combinado).
    items_mo, franquicia = _extraer_mano_obra_y_franquicia(celdas)
    if items_mo:
        resultado["manoObra_items"] = items_mo
    if franquicia:
        resultado["franquicia"] = franquicia

    # Extraer las observaciones del peritaje. Suelen estar en un
    # bloque despues del header "OBSERVACIONES", una linea por
    # observacion.
    observaciones = _extraer_observaciones(celdas)
    if observaciones:
        resultado["observaciones"] = observaciones

    return resultado


def _extraer_observaciones(celdas):
    """
    Busca el header "OBSERVACIONES" en la grilla y recoge las lineas
    de texto que siguen (hasta topar con otra seccion conocida o con
    el final de la grilla). Devuelve las observaciones concatenadas
    con puntos, o "" si no hay.
    """
    # Localizar la fila del header.
    fila_header = None
    col_header = 0
    for (r, c), v in celdas.items():
        if not isinstance(v, str):
            continue
        vu = v.strip().upper()
        if vu == "OBSERVACIONES":
            fila_header = r
            col_header = c
            break
    if fila_header is None:
        return ""

    # Cortar cuando aparezca otra seccion (headers conocidos).
    STOP_HEADERS = (
        "ESTIMACION MANO DE OBRA", "MANO DE OBRA Y REPUESTOS",
        "INSPECTOR ACTUANTE", "RESPONSABLE", "TOTAL", "FRANQUICIA",
        "NETO CIA", "NETO A CARGO",
    )

    lineas = []
    # Buscar filas por debajo del header en la misma columna (o cercanas).
    filas_debajo = sorted(set(r for (r, c) in celdas.keys()
                              if fila_header < r < fila_header + 30))
    for r in filas_debajo:
        # ¿Es alguna otra seccion? -> cortar
        detener = False
        for c in range(20):
            v = celdas.get((r, c))
            if isinstance(v, str):
                vu = v.strip().upper()
                if any(sh in vu for sh in STOP_HEADERS):
                    detener = True
                    break
        if detener:
            break

        # Recolectar los textos de esta fila.
        textos_fila = []
        for c in range(20):
            v = celdas.get((r, c))
            if isinstance(v, str) and v.strip():
                textos_fila.append(v.strip())
        if textos_fila:
            # Filtrar duplicados dentro de la fila (a veces la misma
            # frase aparece en dos columnas por celdas combinadas).
            vistos = []
            for t in textos_fila:
                if t not in vistos:
                    vistos.append(t)
            lineas.append(" ".join(vistos))

    return ". ".join(lineas).strip()


def _extraer_mano_obra_y_franquicia(celdas):
    """
    Extrae los items de mano de obra respetando exactamente lo que
    escribió el perito (no busca por nombre fijo). Ej: si el perito
    combinó "MECANICA Y TAPICERIA" en un solo item, se copia tal cual.

    Devuelve (lista_items, franquicia). Cada item es un dict:
      {"concepto": str, "unidad": str, "cantidad": num, "unitario": num}

    La búsqueda se restringe a partir del header "ESTIMACION MANO DE
    OBRA" para no confundir con items de repuestos.
    """
    # Localizar la fila donde arranca la sección de mano de obra.
    fila_inicio_mo = None
    for (r, c), v in celdas.items():
        if not isinstance(v, str):
            continue
        vu = v.strip().upper()
        if ("MANO DE OBRA" in vu and "ESTIMACI" in vu) or vu == "MANO DE OBRA":
            fila_inicio_mo = r
            break
    if fila_inicio_mo is None:
        return None, 0

    UNIDADES = {"DIAS", "DÍAS", "PAÑOS", "PANOS", "HORAS", "HORA",
                "UNIDAD", "UNIDADES", "PANO", "PAÑO"}
    # Etiquetas que NO son items de mano de obra (encabezados, totales,
    # firmas, franquicia, repuestos).
    NO_ITEMS = {"CANTIDAD", "UNIDAD", "TOTAL", "V.UNITARIO", "V. UNITARIO",
                "VALOR UNITARIO", "SUBTOTAL", "IMPREVISTOS",
                "INSPECTOR", "INSPECTOR ACTUANTE", "RESPONSABLE",
                "FIRMA", "GONZALO", "ESTUDIO", "FERNANDEZ",
                "REPUESTOS", "FRANQUICIA", "NETO", "NETO CIA",
                "NETO CIA.", "MANO DE OBRA", "OBSERVACIONES",
                "ESTIMACION MANO DE OBRA Y REPUESTOS", "OBSERVACION",
                "ITEMS ESTIMADOS"}

    items = []
    franquicia = 0

    # Agrupar celdas por fila para procesarlas en orden.
    filas_ordenadas = sorted(set(r for (r, c) in celdas.keys()
                                 if fila_inicio_mo < r < fila_inicio_mo + 20))

    for r in filas_ordenadas:
        # Etiqueta = primera celda con texto en la fila.
        etiqueta = None
        col_etiqueta = None
        for c in range(20):
            v = celdas.get((r, c))
            if isinstance(v, str) and v.strip():
                etiqueta = v.strip()
                col_etiqueta = c
                break
        if not etiqueta:
            continue

        etiqueta_upper = etiqueta.upper()

        # ¿Es franquicia? La guardamos aparte.
        if "FRANQUICIA" in etiqueta_upper:
            for dc in range(1, 15):
                val = celdas.get((r, col_etiqueta + dc))
                if val is None or val == "":
                    continue
                n = _a_numero(val)
                if n > 0:
                    franquicia = n
                    break
            continue

        # Filtros: NO es un item de mano de obra.
        if etiqueta_upper in NO_ITEMS:
            continue
        if any(nx in etiqueta_upper for nx in
               ("INSPECTOR", "RESPONSABLE", "FIRMA", "ESTUDIO D.A.G",
                "TOTAL MANO", "NETO", "OBSERVACION")):
            continue
        # Muy corto (menos de 4 chars) probablemente no sea concepto real.
        if len(etiqueta) < 4:
            continue

        # Buscar en el resto de la fila: unidad, cantidad, unitario.
        unidad = ""
        cantidad = 0
        unitario = 0
        numeros_encontrados = []  # (columna, valor)
        for dc in range(1, 20):
            val = celdas.get((r, col_etiqueta + dc))
            if val is None or val == "":
                continue
            if isinstance(val, str):
                val_upper = val.strip().upper()
                if val_upper in UNIDADES and not unidad:
                    unidad = val.strip()
                    continue
                n_str = _a_numero(val)
                if n_str > 0:
                    numeros_encontrados.append((col_etiqueta + dc, n_str))
            elif isinstance(val, (int, float)):
                if val > 0:
                    numeros_encontrados.append((col_etiqueta + dc, int(val)))

        # Necesitamos al menos un numero para ser un item valido.
        if not numeros_encontrados:
            continue

        # Heurística: en la estructura típica del peritaje, los números
        # aparecen en este orden: cantidad, unitario, total.
        # - cantidad es "pequeña" (< 1000)
        # - unitario es "grande" (>= 1000, tipicamente 20k-500k)
        # - total es el mayor (cantidad * unitario)
        chicos = [(col, v) for col, v in numeros_encontrados if v < 1000]
        grandes = [(col, v) for col, v in numeros_encontrados if v >= 1000]

        if chicos:
            cantidad = chicos[0][1]
        if grandes:
            # El unitario es el más pequeño de los grandes (el total
            # es cantidad * unitario, así que es más grande).
            unitario = min(v for col, v in grandes)

        # Si no encontramos cantidad chica pero sí uno grande, puede
        # ser un item "monto directo" (ej: solo total en pesos). En
        # ese caso: cantidad=1, unitario=el número.
        if cantidad == 0 and unitario > 0:
            cantidad = 1
        # Si no encontramos unitario pero sí cantidad, dejar unitario
        # en 0 y que se complete a mano.

        items.append({
            "concepto": etiqueta,
            "unidad": unidad,
            "cantidad": cantidad,
            "unitario": unitario,
        })

    # Si no encontramos nada de mano de obra, devolver None.
    if not items:
        return None, franquicia
    return items, franquicia
