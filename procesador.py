"""
procesador.py
Motor de procesamiento de informes de tasacion.

- Usa la API de Claude (Anthropic) para ENTENDER el peritaje en lenguaje
  libre y extraer los datos estructurados, sin importar como esten escritos.
- Completa la plantilla Excel virgen preservando logos, estilos y formulas.

Requiere la variable de entorno ANTHROPIC_API_KEY (configurada en Render).
"""

import os
import re
import json
import zipfile
import unicodedata
from io import BytesIO

from openpyxl import load_workbook
import anthropic


# Modelo de Claude a usar (Sonnet 4.6: buen equilibrio costo/calidad)
MODELO_CLAUDE = "claude-sonnet-4-6"


# ============================================================
#  UTILIDADES
# ============================================================

def normalizar(texto):
    """Pasa a mayusculas y quita tildes."""
    if texto is None:
        return ""
    texto = str(texto).upper()
    texto = unicodedata.normalize("NFD", texto)
    texto = "".join(c for c in texto if unicodedata.category(c) != "Mn")
    return texto.strip()


def a_numero(valor):
    """Convierte un string a numero, quitando $, puntos y comas."""
    if valor is None:
        return 0
    s = re.sub(r"[^\d]", "", str(valor))
    return int(s) if s else 0


# ============================================================
#  ESTRUCTURA DE DATOS
# ============================================================

def data_vacia():
    return {
        "numeroSiniestro": "", "fechaSiniestro": "", "fechaInspeccion": "",
        "asegurado": "", "marca": "", "modelo": "", "anio": "", "dominio": "",
        "chasis": "", "kilometraje": "", "sumaAsegurada": "", "franquiciaVeh": "",
        "tallerNombre": "", "tallerDireccion": "", "tallerLocalidad": "",
        "danos": [],  # lista de {accion, pieza, precio}
        "manoObra": {
            "pintura": 0, "chapa": 0, "mecanica": 0, "tapiceria": 0, "varios": 0,
            "pinturaValor": 120000, "chapaValor": 120000,
            "mecanicaValor": 50000, "tapiceriaValor": 50000,
        },
        "franquicia": 0,
        "observaciones": "",
    }


# ============================================================
#  CEREBRO: ENTENDER EL PERITAJE CON LA API DE CLAUDE
# ============================================================

# Instrucciones para Claude: como perito, que datos extraer y en que formato.
_PROMPT_SISTEMA_PARSEO = """Sos un perito tasador de seguros automotores con \
experiencia. Recibis el texto libre de una peritacion (puede venir \
desordenado, con abreviaturas, datos de poliza mezclados, etc.) y tu tarea \
es EXTRAER los datos y devolverlos en formato JSON estricto.

Reglas de interpretacion:
- Las piezas bajo "SUSTITUIR" o "CAMBIAR" tienen accion "CAMBIAR".
- Las piezas bajo "PINTAR" o "REPARAR" tienen accion "REPARAR".

- PERITAJE DESDE EXCEL (GRILLA): si el texto incluye un bloque que
  empieza con "=== PERITAJE DESDE EXCEL (GRILLA) ===", ese bloque es el
  volcado de una planilla Excel de peritacion, celda por celda, con sus
  coordenadas (ej: "A55=Paragolpe | B55=X | D55=95000"). Interpretala
  asi:
  * La planilla tiene una grilla de danos dividida en SECTORES. Cada
    sector tiene un encabezado de texto (ej: "PARTE DELANTERA",
    "PARTE TRASERA", "LADO IZQUIERDO", "LADO DERECHO", "PARTE INTERIOR",
    "MOTOR", "CHASIS", "TREN TRASERO", "TREN DELANTERO", "OTROS").
  * Cada sector tiene un par de columnas marcadas "A" y "B" en su fila
    de encabezado, y una columna "Precio". Las piezas son los textos
    que estan a la izquierda de esas columnas.
  * Una "X" en la columna "A" del sector significa que esa pieza va a
    CAMBIAR. Una "X" en la columna "B" significa REPARAR.
  * Recorré TODA la grilla y detectá TODAS las "X", en todos los
    sectores (las columnas A/B de cada sector estan en distintas
    posiciones: pueden ser B/C, G/H, L/M, etc.).
  * El nombre de cada pieza se arma combinando el SECTOR donde esta con
    el texto de la fila, para que quede sin ambiguedad. Ejemplos:
    "Paragolpe" en el sector "PARTE TRASERA" -> "Paragolpe trasero".
    "Guardabarro der." en "PARTE TRASERA" -> "Guardabarro trasero
    derecho". "Guardabarro der." en "PARTE DELANTERA" -> "Guardabarro
    delantero derecho".
  * IGNORÁ los precios que vengan en la columna "Precio" del Excel. Los
    precios NO se toman del Excel del peritaje.
  * Del Excel tambien extraé los datos del vehiculo, asegurado, taller,
    mano de obra y observaciones si estan presentes.

- PRECIOS DE REPUESTOS: el texto puede incluir, despues de una linea
  "=== COTIZACION DE REPUESTOS ===", una tabla o lista de precios de
  repuestos (con columnas tipo Repuesto, Precio s/IVA, Precio c/IVA,
  etc.). Si esa cotizacion esta presente:
  * Para cada pieza a CAMBIAR, asignale su precio usando SIEMPRE la
    columna "Precio c/IVA" (el valor CON IVA).
  * Emparejá cada repuesto de la cotizacion con la pieza del peritaje
    aunque esten escritos distinto. Ejemplos: "PPE Del" del peritaje =
    "Paragolpe delantero" de la cotizacion; "Felpa Bajo Capot" =
    "Felpa / manta bajo capot"; "Optica Izq" = "Optica delantera
    izquierda".
  * El precio es solo el numero entero, sin signo $ ni puntos de miles.
  * Si una pieza a CAMBIAR no aparece en la cotizacion, su precio es 0.
- Tambien se acepta un precio escrito al lado de la pieza en el peritaje
  (ej: "Capot $180000").
- Si NO hay ninguna cotizacion, todas las piezas a CAMBIAR van con
  precio 0.
- Las piezas a REPARAR siempre llevan precio 0.
- Mano de obra: interpretar todas las modalidades de escritura. Ejemplos:
  "7 panos", "pint 7", "7p" -> pintura = 7.
  "chapa 3", "3 dias", "3d", "3 jornadas" -> chapa = 3.
  "12 hs mecanica", "mecanica 12" -> mecanica = 12.
- "carga de gas", "varios", "service" -> sumar al campo manoObra.varios.
- El numero de siniestro puede venir solo (un numero suelto) o con etiqueta.
- La fecha de inspeccion puede venir como "fecha ip 16/05/26", "16-05-26", \
etc., en cualquier parte del texto. En el Excel suele ser "Dia de
Inspeccion"; si viene como numero serial de Excel, convertilo a fecha.
- La suma asegurada es solo el numero, sin texto pegado.
- Los datos del taller / lugar de inspeccion (nombre, direccion, localidad) \
pueden aparecer en cualquier parte; extraelos si los encontras.
- Las observaciones son la descripcion en prosa del perito (ej: "visto en \
domicilio particular..."), NO los datos de poliza, productor, vigencia, etc.
- Si un dato no aparece, dejalo como cadena vacia "" o 0 segun corresponda.

Devolve UNICAMENTE un objeto JSON valido, sin texto antes ni despues, sin \
backticks, con esta estructura exacta:
{
  "numeroSiniestro": "",
  "fechaSiniestro": "",
  "fechaInspeccion": "",
  "asegurado": "",
  "marca": "",
  "modelo": "",
  "anio": "",
  "dominio": "",
  "chasis": "",
  "kilometraje": "",
  "sumaAsegurada": "",
  "tallerNombre": "",
  "tallerDireccion": "",
  "tallerLocalidad": "",
  "franquicia": 0,
  "manoObra": {"pintura": 0, "chapa": 0, "mecanica": 0,
               "tapiceria": 0, "varios": 0},
  "danos": [{"accion": "CAMBIAR", "pieza": "nombre", "precio": 0}],
  "observaciones": ""
}"""


def _cliente_claude():
    """Crea el cliente de la API de Claude. Lee la key del entorno."""
    api_key = os.environ.get("ANTHROPIC_API_KEY", "").strip()
    if not api_key:
        raise RuntimeError(
            "Falta la variable de entorno ANTHROPIC_API_KEY. "
            "Configurala en Render (Environment).")
    return anthropic.Anthropic(api_key=api_key)


def _extraer_json(texto):
    """Extrae el primer objeto JSON de un texto, tolerando backticks."""
    t = texto.strip()
    # Quitar fences ```json ... ```
    t = re.sub(r"^```(?:json)?\s*|\s*```$", "", t,
               flags=re.IGNORECASE | re.MULTILINE)
    # Tomar desde la primera { hasta la ultima }
    ini = t.find("{")
    fin = t.rfind("}")
    if ini >= 0 and fin > ini:
        t = t[ini:fin + 1]
    return json.loads(t)


def parsear_texto(texto):
    """
    Entiende el texto libre de la peritacion usando la API de Claude
    y devuelve el dict de datos estructurado.

    Si Claude falla o el texto esta vacio, devuelve data_vacia().
    """
    data = data_vacia()
    if not texto or not texto.strip():
        return data

    cliente = _cliente_claude()
    respuesta = cliente.messages.create(
        model=MODELO_CLAUDE,
        max_tokens=8000,
        system=_PROMPT_SISTEMA_PARSEO,
        messages=[{
            "role": "user",
            "content": "Texto de la peritacion:\n\n" + texto,
        }],
    )

    # Juntar el texto devuelto por Claude
    salida = ""
    for bloque in respuesta.content:
        if getattr(bloque, "type", None) == "text":
            salida += bloque.text

    try:
        parsed = _extraer_json(salida)
    except Exception:
        # Si Claude no devolvio JSON valido, no se pierde el informe:
        # se devuelve data vacia y el resto del flujo sigue.
        return data

    # Volcar los campos al dict, con valores por defecto seguros
    for campo in ("numeroSiniestro", "fechaSiniestro", "fechaInspeccion",
                  "asegurado", "marca", "modelo", "anio", "dominio",
                  "chasis", "kilometraje", "sumaAsegurada",
                  "tallerNombre", "tallerDireccion", "tallerLocalidad",
                  "observaciones"):
        v = parsed.get(campo, "")
        data[campo] = str(v).strip() if v is not None else ""

    data["franquicia"] = a_numero(parsed.get("franquicia", 0))

    mo = parsed.get("manoObra", {}) or {}
    for k in ("pintura", "chapa", "mecanica", "tapiceria", "varios"):
        data["manoObra"][k] = a_numero(mo.get(k, 0))

    danos = []
    for d in parsed.get("danos", []) or []:
        accion = normalizar(d.get("accion", ""))
        accion = "CAMBIAR" if accion not in ("REPARAR",) else "REPARAR"
        pieza = str(d.get("pieza", "")).strip()
        # Solo las piezas a CAMBIAR llevan precio; las REPARAR van en 0.
        precio = a_numero(d.get("precio", 0)) if accion == "CAMBIAR" else 0
        if pieza:
            danos.append({"accion": accion, "pieza": pieza,
                          "precio": precio})
    data["danos"] = danos

    return data
# ============================================================
#  PRESERVACION DE IMAGENES (LOGOS)
# ============================================================

def _reinyectar_imagenes(plantilla_bytes, generado_bytes):
    """
    Restaura los logos de la plantilla en el informe generado, PERO solo
    si openpyxl los perdio al guardar.

    En algunos entornos (segun version de openpyxl / del sistema), al
    hacer load_workbook + save las imagenes ancladas se descartan.
    Esta funcion lo detecta y, si faltan, reinyecta a nivel ZIP:
      - xl/media/*      (las imagenes)
      - xl/drawings/*   (los anclajes y sus rels)
    y ademas, para cada hoja que tenia imagenes:
      - agrega la relacion al drawing en su xl/worksheets/_rels/sheetN.xml.rels
      - inserta la etiqueta <drawing r:id="..."/> dentro de sheetN.xml
        (sin esto, Excel NO muestra la imagen aunque el archivo este)
    y declara los drawings en [Content_Types].xml.

    Si openpyxl ya conservo las imagenes, NO toca nada.
    Si la plantilla no tiene imagenes, tampoco actua.
    """
    try:
        zp = zipfile.ZipFile(BytesIO(plantilla_bytes))
        zg = zipfile.ZipFile(BytesIO(generado_bytes))
    except zipfile.BadZipFile:
        return generado_bytes

    nombres_p = set(zp.namelist())
    nombres_g = set(zg.namelist())

    # La plantilla no tiene imagenes -> nada que reinyectar
    if not any(n.startswith("xl/media/") for n in nombres_p):
        return generado_bytes

    # El generado YA conservo las imagenes -> no tocar
    if any(n.startswith("xl/media/") for n in nombres_g):
        return generado_bytes

    # --- Mapear: cada sheetN -> que drawing usa (segun la plantilla) ---
    # En la plantilla, xl/worksheets/_rels/sheetN.xml.rels apunta al drawing.
    sheet_a_drawing = {}  # 'sheet1' -> 'drawing1'
    for n in nombres_p:
        m = re.match(r"xl/worksheets/_rels/(sheet\d+)\.xml\.rels$", n)
        if m:
            contenido = zp.read(n).decode("utf-8")
            md = re.search(r'Target="\.\./drawings/(drawing\d+\.xml)"',
                           contenido)
            if md:
                sheet_a_drawing[m.group(1)] = md.group(1)

    # Archivos de imagen a copiar tal cual de la plantilla
    grupo = [n for n in nombres_p
             if n.startswith("xl/media/") or n.startswith("xl/drawings/")]

    # --- Fusionar [Content_Types].xml ---
    ct = zg.read("[Content_Types].xml").decode("utf-8")
    extras = ""
    for dw in sorted(set(sheet_a_drawing.values())):
        part = "/xl/drawings/" + dw
        if part not in ct:
            extras += ('<Override PartName="%s" ContentType='
                       '"application/vnd.openxmlformats-officedocument'
                       '.drawing+xml"/>' % part)
    if 'Extension="png"' not in ct:
        extras = ('<Default Extension="png" ContentType="image/png"/>'
                  + extras)
    if extras:
        ct = ct.replace("</Types>", extras + "</Types>")

    def _agregar_drawing_a_rels(rels_xml, drawing_file):
        """Garantiza una relacion al drawing y devuelve (xml_nuevo, rId).
        Si ya hay una relacion de tipo drawing, reutiliza su Id."""
        # Si ya existe una relacion de drawing, usar ese rId
        m = re.search(
            r'<Relationship[^>]*relationships/drawing"[^>]*Id="(rId\d+)"',
            rels_xml)
        if not m:
            m = re.search(
                r'<Relationship[^>]*Id="(rId\d+)"[^>]*relationships/drawing"',
                rels_xml)
        if m:
            rid = m.group(1)
            # Reescribir esa relacion para que apunte al drawing correcto
            rels_xml = re.sub(
                r'<Relationship[^>]*relationships/drawing"[^>]*/>',
                '', rels_xml)
            rels_xml = re.sub(
                r'<Relationship[^>]*Id="' + rid + r'"[^>]*/>',
                '', rels_xml)
        else:
            usados = re.findall(r'Id="rId(\d+)"', rels_xml)
            n = max([int(x) for x in usados], default=0) + 1
            rid = "rId%d" % n
        rel = ('<Relationship Id="%s" Type="http://schemas.openxmlformats'
               '.org/officeDocument/2006/relationships/drawing" '
               'Target="../drawings/%s"/>' % (rid, drawing_file))
        nuevo = rels_xml.replace("</Relationships>",
                                 rel + "</Relationships>")
        return nuevo, rid

    def _insertar_drawing_en_sheet(sheet_xml, rid):
        """Garantiza <drawing r:id=.../> antes de </worksheet>.
        Si ya hay una etiqueta <drawing>, la reemplaza por la correcta.
        Tambien garantiza que el namespace 'r' este declarado, ya que
        sin xmlns:r el atributo r:id provoca 'unbound prefix' en Excel."""
        # Quitar cualquier <drawing.../> existente (puede estar roto)
        sheet_xml = re.sub(r'<drawing\b[^>]*/>', '', sheet_xml)
        # Asegurar que el tag raiz <worksheet ...> declare xmlns:r
        if "xmlns:r=" not in sheet_xml:
            ns = ('xmlns:r="http://schemas.openxmlformats.org/'
                  'officeDocument/2006/relationships"')
            sheet_xml = re.sub(
                r'(<worksheet\b)', r'\1 ' + ns, sheet_xml, count=1)
        tag = '<drawing r:id="%s"/>' % rid
        return sheet_xml.replace("</worksheet>", tag + "</worksheet>")

    out = BytesIO()
    with zipfile.ZipFile(out, "w", zipfile.ZIP_DEFLATED) as zout:
        escritos = set()

        for item in zg.namelist():
            # Content_Types fusionado
            if item == "[Content_Types].xml":
                zout.writestr(item, ct)
                escritos.add(item)
                continue

            # Los .rels de hojas se manejan junto a su sheet -> saltar aqui
            if re.match(r"xl/worksheets/_rels/sheet\d+\.xml\.rels$", item):
                continue

            # sheetN.xml: insertar la etiqueta <drawing>
            ms = re.match(r"xl/worksheets/(sheet\d+)\.xml$", item)
            if ms and ms.group(1) in sheet_a_drawing:
                sheet_id = ms.group(1)
                dw = sheet_a_drawing[sheet_id]
                rels_path = ("xl/worksheets/_rels/%s.xml.rels"
                             % sheet_id)
                if rels_path in nombres_g:
                    rels_xml = zg.read(rels_path).decode("utf-8")
                else:
                    rels_xml = ('<?xml version="1.0" encoding="UTF-8" '
                                'standalone="yes"?>'
                                '<Relationships xmlns="http://schemas'
                                '.openxmlformats.org/package/2006/'
                                'relationships"></Relationships>')
                rels_nuevo, rid = _agregar_drawing_a_rels(rels_xml, dw)
                sheet_xml = zg.read(item).decode("utf-8")
                sheet_xml = _insertar_drawing_en_sheet(sheet_xml, rid)
                zout.writestr(item, sheet_xml)
                zout.writestr(rels_path, rels_nuevo)
                escritos.add(item)
                escritos.add(rels_path)
                continue

            zout.writestr(item, zg.read(item))
            escritos.add(item)

        # Copiar imagenes y drawings de la plantilla
        for item in grupo:
            if item not in escritos:
                zout.writestr(item, zp.read(item))
                escritos.add(item)

    return out.getvalue()



# ============================================================
#  COMPLETAR LA PLANTILLA (PRESERVA LOGOS Y FORMATO)
# ============================================================

def _rango_de(ws, fila, columna):
    """Devuelve el rango combinado que contiene a (fila, columna), o None."""
    for rango in ws.merged_cells.ranges:
        if (rango.min_row <= fila <= rango.max_row and
                rango.min_col <= columna <= rango.max_col):
            return rango
    return None


def _valor_celda(ws, fila, columna):
    """
    Lee el valor de una celda de forma segura.
    Si es parte de un rango combinado, el valor real esta en la celda
    ancla (esquina superior izquierda).
    """
    rango = _rango_de(ws, fila, columna)
    if rango is not None:
        return ws.cell(row=rango.min_row, column=rango.min_col).value
    return ws.cell(row=fila, column=columna).value


def _escribir(ws, fila, columna, valor):
    """
    Escribe un valor en una celda manejando celdas combinadas.

    FIX del error "'MergedCell' object attribute 'value' is read-only":
    si la celda destino cae dentro de un rango combinado, se DESHACE el
    merge antes de escribir y se vuelve a combinar despues. Asi la celda
    deja de ser una MergedCell inmutable.
    """
    rango = _rango_de(ws, fila, columna)
    if rango is not None:
        rango_str = str(rango)
        anc_row, anc_col = rango.min_row, rango.min_col
        ws.unmerge_cells(rango_str)
        ws.cell(row=anc_row, column=anc_col).value = valor
        ws.merge_cells(rango_str)
    else:
        ws.cell(row=fila, column=columna).value = valor


def _es_celda_etiqueta(valor_celda, buscado):
    """
    Determina si una celda ES una etiqueta de campo (no que la contiene
    como substring accidental, ej. 'ANO' dentro de 'EMILIANO').

    El texto buscado debe aparecer como PALABRA COMPLETA dentro de una
    celda que se comporte como etiqueta (texto corto o terminado en ':').
    """
    v = normalizar(valor_celda)
    b = normalizar(buscado)
    if not v or not b:
        return False
    # 'b' como palabra completa dentro de 'v'
    if not re.search(r"(?<![A-Z0-9])" + re.escape(b) + r"(?![A-Z0-9])", v):
        return False
    # La celda debe parecer etiqueta: termina en ':'/'=' o es texto corto
    v_limpio = v.rstrip(": =").strip()
    if v.endswith(":") or v.endswith("=") or len(v_limpio) <= 35:
        return True
    return False


def _buscar_celda_etiqueta(ws, textos_buscados, ocurrencia=1):
    """
    Busca la celda que ES una etiqueta y devuelve la coordenada (fila, columna)
    de la celda donde escribir el valor (la primera celda vacia a la derecha).
    Devuelve None si no encuentra la etiqueta.
    """
    encontradas = 0
    for fila_celdas in ws.iter_rows():
        for celda in fila_celdas:
            if not normalizar(celda.value):
                continue
            for buscado in textos_buscados:
                if _es_celda_etiqueta(celda.value, buscado):
                    encontradas += 1
                    if encontradas == ocurrencia:
                        fila, col = celda.row, celda.column
                        for c in range(col + 1, col + 9):
                            if _valor_celda(ws, fila, c) in (None, ""):
                                return (fila, c)
                        return (fila, col + 1)
    return None


def completar_plantilla(plantilla_bytes, data):
    """
    Carga la plantilla virgen, la completa con los datos y devuelve los bytes.
    Preserva logos/imagenes, estilos, formulas y formato de impresion.
    Maneja correctamente las celdas combinadas (merged cells).
    """
    wb = load_workbook(BytesIO(plantilla_bytes))
    hojas = wb.worksheets

    # ---------- HOJA 1: Datos preliminares ----------
    ws1 = hojas[0]

    mapeo_hoja1 = [
        (["NUMERO DE SINIESTRO", "NRO SINIESTRO"], data["numeroSiniestro"]),
        (["FECHA DE SINIESTRO", "FECHA SINIESTRO"], data["fechaSiniestro"]),
        (["FECHA DE INSPECCION", "FECHA INSPECCION"], data["fechaInspeccion"]),
        (["APELLIDO Y NOMBRE"], data["asegurado"]),
        (["MARCA"], data["marca"]),
        (["MODELO"], data["modelo"]),
        (["ANO", "ANIO"], data["anio"]),
        (["DOMINIO", "PATENTE"], data["dominio"]),
        (["CHASIS"], data["chasis"]),
        (["KILOMETRAJE", "KILOMETRA"], data["kilometraje"]),
        (["SUMA ASEGURADA", "SUMA ASEG"], data["sumaAsegurada"]),
    ]
    for etiquetas, valor in mapeo_hoja1:
        if not valor:
            continue
        pos = _buscar_celda_etiqueta(ws1, etiquetas)
        if pos is not None:
            _escribir(ws1, pos[0], pos[1], valor)

    # ---- Datos del taller / lugar de inspeccion ----
    # Se busca primero la fila de la seccion "DATOS DEL TALLER" y, a
    # partir de ahi, las etiquetas NOMBRE/DIRECCION/LOCALIDAD. Asi no se
    # confunde "NOMBRE" con el "APELLIDO Y NOMBRE" del asegurado.
    fila_taller = None
    for fila_celdas in ws1.iter_rows():
        for celda in fila_celdas:
            v = normalizar(celda.value)
            if "DATOS DEL TALLER" in v or "LUGAR DE INSPECCION" in v:
                fila_taller = celda.row
                break
        if fila_taller:
            break

    if fila_taller is not None:
        taller_map = [
            (["NOMBRE"], data.get("tallerNombre", "")),
            (["DIRECCION"], data.get("tallerDireccion", "")),
            (["LOCALIDAD"], data.get("tallerLocalidad", "")),
        ]
        for etiquetas, valor in taller_map:
            if not valor:
                continue
            # Buscar la etiqueta en las filas de la seccion del taller.
            # El valor va en la primera celda vacia a la DERECHA de la
            # etiqueta (saltando el ancho de celdas combinadas).
            for fila_celdas in ws1.iter_rows(min_row=fila_taller,
                                             max_row=fila_taller + 6):
                encontrada = False
                for celda in fila_celdas:
                    et_ok = any(_es_celda_etiqueta(celda.value, et)
                                for et in etiquetas)
                    if not et_ok:
                        continue
                    # Punto de partida: columna siguiente al merge de
                    # la etiqueta (o a la celda si no esta combinada).
                    rango = _rango_de(ws1, celda.row, celda.column)
                    col_ini = (rango.max_col + 1) if rango \
                        else (celda.column + 1)
                    destino = None
                    for c in range(col_ini, col_ini + 9):
                        if _valor_celda(ws1, celda.row, c) in (None, ""):
                            destino = c
                            break
                    if destino is None:
                        destino = col_ini
                    _escribir(ws1, celda.row, destino, valor)
                    encontrada = True
                    break
                if encontrada:
                    break

    # Franquicia del vehiculo (1ra ocurrencia de "FRANQUICIA").
    # Regla: si no hay valor, se escribe 0 (no se deja vacio).
    franq_veh = a_numero(data.get("franquiciaVeh")) if data.get("franquiciaVeh") else 0
    pos = _buscar_celda_etiqueta(ws1, ["FRANQUICIA"], ocurrencia=1)
    if pos is not None:
        _escribir(ws1, pos[0], pos[1], franq_veh)

    # Franquicia a deducir (2da ocurrencia, "FRANQUICIA:").
    # La Hoja3 la lee desde aqui con una formula. Si no hay valor, va 0.
    franq_ded = a_numero(data.get("franquicia")) if data.get("franquicia") else 0
    pos = _buscar_celda_etiqueta(ws1, ["FRANQUICIA"], ocurrencia=2)
    if pos is not None:
        _escribir(ws1, pos[0], pos[1], franq_ded)

    # ---------- HOJA 2: Descripcion de danos ----------
    if len(hojas) > 1:
        ws2 = hojas[1]
        fila_inicio = None
        col_accion = col_pieza = col_precio = None
        for fila_celdas in ws2.iter_rows():
            for celda in fila_celdas:
                v = normalizar(celda.value)
                if v in ("ACCION", "ACCION:"):
                    fila_inicio = celda.row + 1
                    col_accion = celda.column
                if "PIEZA" in v:
                    col_pieza = celda.column
                if "PRECIO" in v:
                    col_precio = celda.column
            if fila_inicio:
                break

        if fila_inicio:
            col_accion = col_accion or 1
            col_pieza = col_pieza or 2
            col_precio = col_precio or 6
            for i, dano in enumerate(data["danos"]):
                fila = fila_inicio + i
                _escribir(ws2, fila, col_accion, dano["accion"])
                _escribir(ws2, fila, col_pieza, dano["pieza"])
                # Solo las piezas a CAMBIAR llevan precio.
                # Si la pieza es CAMBIAR pero no tiene precio, va 0.
                if dano["accion"] == "CAMBIAR":
                    _escribir(ws2, fila, col_precio, dano.get("precio", 0) or 0)

    # ---------- HOJA 3: Mano de obra, resumen, observaciones ----------
    if len(hojas) > 2:
        ws3 = hojas[2]
        mo = data["manoObra"]

        # ---- Total Repuestos ----
        # La celda "Total Repuestos" de la Hoja3 NO es una formula: hay
        # que escribir la suma de los precios de las piezas a CAMBIAR.
        total_repuestos = sum(
            (d.get("precio", 0) or 0)
            for d in data["danos"]
            if d["accion"] == "CAMBIAR"
        )
        for fila_celdas in ws3.iter_rows():
            encontrada = False
            for celda in fila_celdas:
                if "TOTAL REPUESTOS" in normalizar(celda.value):
                    # El valor va en la celda de importe de la misma
                    # fila (a la derecha de la etiqueta).
                    rango = _rango_de(ws3, celda.row, celda.column)
                    col_valor = (rango.max_col + 1) if rango \
                        else (celda.column + 1)
                    _escribir(ws3, celda.row, col_valor, total_repuestos)
                    encontrada = True
                    break
            if encontrada:
                break

        def _buscar_cols_cant_vu(fila_concepto):
            """Busca las columnas CANT. y V. UNITARIO mirando filas de arriba."""
            col_cant = col_vu = None
            for r in range(max(1, fila_concepto - 6), fila_concepto):
                for c in range(1, 12):
                    hv = normalizar(_valor_celda(ws3, r, c))
                    if "CANT" in hv:
                        col_cant = c
                    if "UNITARIO" in hv or "V. UNIT" in hv:
                        col_vu = c
            return col_cant, col_vu

        def set_mano_obra(nombre_concepto, cantidad, valor_unitario):
            """Busca la fila del concepto y carga cantidad y valor unitario."""
            objetivo = normalizar(nombre_concepto)
            for fila_celdas in ws3.iter_rows():
                for celda in fila_celdas:
                    if normalizar(celda.value) == objetivo:
                        col_cant, col_vu = _buscar_cols_cant_vu(celda.row)
                        if col_cant:
                            _escribir(ws3, celda.row, col_cant, cantidad)
                        if col_vu and valor_unitario:
                            _escribir(ws3, celda.row, col_vu, valor_unitario)
                        return

        set_mano_obra("Pintura", mo["pintura"], mo["pinturaValor"])
        set_mano_obra("Chapa", mo["chapa"], mo["chapaValor"])
        set_mano_obra("Mecanica", mo["mecanica"], mo["mecanicaValor"])
        set_mano_obra("Tapiceria", mo["tapiceria"], mo["tapiceriaValor"])

        # Varios: la fila "Varios" tiene formula SUBTOTAL = CANT * V.UNITARIO.
        # Para no pisar la formula, se carga el monto como V.UNITARIO con CANT=1.
        if mo["varios"]:
            for fila_celdas in ws3.iter_rows():
                for celda in fila_celdas:
                    if normalizar(celda.value) == "VARIOS":
                        col_cant, col_vu = _buscar_cols_cant_vu(celda.row)
                        if col_cant:
                            _escribir(ws3, celda.row, col_cant, 1)
                        if col_vu:
                            _escribir(ws3, celda.row, col_vu, mo["varios"])

        # Franquicia a deducir: NO se escribe si la celda tiene formula
        # (suele ser =+Hoja1!$F$28). El valor real se carga en Hoja1.
        if data.get("franquicia"):
            for fila_celdas in ws3.iter_rows():
                for celda in fila_celdas:
                    if "FRANQUICIA A DEDUCIR" in normalizar(celda.value):
                        for c in range(celda.column + 1, celda.column + 9):
                            val = _valor_celda(ws3, celda.row, c)
                            if isinstance(val, str) and val.startswith("="):
                                continue
                            if val in (None, ""):
                                _escribir(ws3, celda.row, c, data["franquicia"])
                                break

        # Observaciones: la etiqueta "OBSERVACIONES" suele estar en una celda
        # combinada. El texto va en el area (merge) inmediatamente debajo,
        # NUNCA sobre la etiqueta misma.
        if data.get("observaciones"):
            for fila_celdas in ws3.iter_rows():
                hecho = False
                for celda in fila_celdas:
                    if normalizar(celda.value) == "OBSERVACIONES":
                        # Si la etiqueta esta en un merge, escribir en la
                        # primera fila despues de ese rango combinado.
                        rango = _rango_de(ws3, celda.row, celda.column)
                        if rango is not None:
                            fila_dest = rango.max_row + 1
                        else:
                            fila_dest = celda.row + 1
                        _escribir(ws3, fila_dest, celda.column,
                                  data["observaciones"])
                        hecho = True
                        break
                if hecho:
                    break

    # Guardar a bytes.
    salida = BytesIO()
    wb.save(salida)
    salida.seek(0)
    generado = salida.read()

    # PRESERVAR LOGOS: si openpyxl perdio las imagenes al guardar
    # (ocurre en algunos entornos), se reinyectan desde la plantilla.
    # Si openpyxl ya las conservo, esta funcion no hace nada.
    return _reinyectar_imagenes(plantilla_bytes, generado)



# ============================================================
#  FUNCION PRINCIPAL
# ============================================================

def procesar(plantilla_bytes, texto_peritacion="", excel_peritacion_bytes=None,
             cotizar=False):
    """
    Funcion principal: recibe la plantilla y la peritacion,
    devuelve (informe_bytes, data, detalle_cotizacion).

    - texto_peritacion: texto libre de la peritacion (lo entiende Claude).
    - excel_peritacion_bytes: se acepta por compatibilidad; reservado.
    - cotizar: parametro mantenido por compatibilidad. La cotizacion
      automatica fue descartada; los precios se cargan manualmente.

    detalle_cotizacion siempre se devuelve como lista vacia.
    """
    data = parsear_texto(texto_peritacion) if texto_peritacion else data_vacia()
    informe_bytes = completar_plantilla(plantilla_bytes, data)
    return informe_bytes, data, []
